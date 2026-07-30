#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
独立数据库更新模块（改造版）
- 读取 Duckdb数据库结构.xlsx 动态建表，仅保留“是否写入”为“是”的字段
- 支持全量/增量更新，自动检测表结构变更并同步
- 多Token轮询 + 无限等待，不遗漏任何交易日
- 线程池并发获取日度历史数据（仅使用 daily 接口）
- 当日信息表使用高级Token获取 daily_basic 数据
- 自动计算周度/月度聚合表
- 健壮的重试机制：所有失败日期最多重试5轮，最终降级为单线程间隔1.5秒
"""

import os
import time
import threading
import logging
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import duckdb
import pandas as pd
import tushare as ts
from openpyxl import load_workbook

# ---------- 全局配置 ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, 'instance')
DB_PATH = os.path.join(INSTANCE_DIR, 'stock_data.duckdb')
EXCEL_PATH = os.path.join(BASE_DIR, 'Duckdb数据库结构.xlsx')

# Tushare Tokens（普通积分，用于 daily / stock_basic 等）
TUSHARE_TOKENS = [
    '2288acdb7397a302d5b7160a2d5878eebf95db0c774d5c3e042918cc',
    'e8bc8974b028e1e7a54fb844320ec248c93025fd1d2bd5eee79c5232',
    '285e90f673a28841080220492009989b613dce0a622edc22f28f9871',
    'b5f46fa9591ab477b028f70c218fb12be223d580cb53afd449dfc216',
    'e9bac38172fdb42c76a46b58e25ffefa747cfb53dffed5ef74e93240'
]

# Tushare Tokens（2000积分，用于 daily_basic 接口）
TUSHARE_TOKENS_2000__ = ['2288acdb7397a302d5b7160a2d5878eebf95db0c774d5c3e042918cc']

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# ---------- 全局缓存 ----------
_DB_STRUCTURE_CACHE = None          # 缓存解析后的结构（仅保留“是”的字段）
_DB_STRUCTURE_RAW_CACHE = None      # 原始结构（含所有字段，用于对比）

# ---------- 线程本地存储 ----------
_thread_local = threading.local()

def get_thread_db_connection():
    if not hasattr(_thread_local, 'conn') or _thread_local.conn is None:
        _thread_local.conn = duckdb.connect(DB_PATH)
    return _thread_local.conn

# ---------- 辅助函数 ----------
def get_latest_trade_date_with_data(pro, max_retry=5):
    today = datetime.now().strftime('%Y%m%d')
    candidate = today
    for _ in range(max_retry):
        try:
            df_test = pro.daily(trade_date=candidate, limit=1)
            if df_test is not None and not df_test.empty:
                return candidate
        except:
            pass
        try:
            cal = pro.trade_cal(exchange='SSE', start_date=candidate, end_date=candidate, is_open='1')
            if cal is not None and not cal.empty:
                prev_date = (datetime.strptime(candidate, '%Y%m%d') - timedelta(days=1)).strftime('%Y%m%d')
                cal_prev = pro.trade_cal(exchange='SSE', start_date=prev_date, end_date=prev_date, is_open='1')
                if cal_prev is not None and not cal_prev.empty:
                    candidate = prev_date
                else:
                    cal_all = pro.trade_cal(exchange='SSE',
                                            start_date=(datetime.now() - timedelta(days=10)).strftime('%Y%m%d'),
                                            end_date=candidate, is_open='1')
                    if cal_all is not None and not cal_all.empty:
                        sorted_dates = sorted(cal_all['cal_date'].tolist())
                        if len(sorted_dates) >= 2:
                            candidate = sorted_dates[-2]
                        elif len(sorted_dates) == 1:
                            candidate = sorted_dates[-1]
                        else:
                            raise Exception("无法确定最近有数据的交易日：近10日交易日历为空")
                    else:
                        raise Exception("无法确定最近有数据的交易日：近10日无开市记录")
            else:
                cal_all = pro.trade_cal(exchange='SSE',
                                        start_date=(datetime.now() - timedelta(days=10)).strftime('%Y%m%d'),
                                        end_date=candidate, is_open='1')
                if cal_all is not None and not cal_all.empty:
                    sorted_dates = sorted(cal_all['cal_date'].tolist())
                    if sorted_dates:
                        candidate = sorted_dates[-1]
                    else:
                        raise Exception("无法确定最近有数据的交易日：候选日非交易日且近10日无开市记录")
                else:
                    raise Exception("无法确定最近有数据的交易日：候选日非交易日且近10日交易日历查询失败")
        except Exception:
            raise
    raise Exception(f"无法确定最近有数据的交易日：重试 {max_retry} 轮后仍未找到（最后候选 {candidate}）")

def get_trade_calendar(pro, start_date=None, end_date=None):
    if start_date is None:
        start_date = '19900101'
    if end_date is None:
        end_date = datetime.now().strftime('%Y%m%d')
    df = pro.trade_cal(exchange='SSE', start_date=start_date, end_date=end_date, is_open='1')
    if df is None or df.empty:
        return []
    return sorted(df['cal_date'].tolist())

# ---------- Token 管理器 ----------
class TokenManager:
    def __init__(self, tokens):
        self.tokens = tokens
        self.lock = threading.Lock()
        self.used_tokens = set()

    def acquire_token(self):
        while True:
            with self.lock:
                for token in self.tokens:
                    if token not in self.used_tokens:
                        self.used_tokens.add(token)
                        return RateLimitedToken(token, self)
            time.sleep(0.5)

    def release_token(self, token):
        with self.lock:
            self.used_tokens.discard(token)

class RateLimitedToken:
    def __init__(self, token, manager):
        self.token = token
        self.manager = manager
        self.last_request_time = 0
        self.lock = threading.Lock()
        self.interval = 1.2  # 1.2秒间隔，确保50次/分钟以内

    def request(self, func, *args, **kwargs):
        with self.lock:
            now = time.time()
            wait = self.interval - (now - self.last_request_time)
            if wait > 0:
                time.sleep(wait)
            ts.set_token(self.token)
            pro = ts.pro_api()
            result = func(pro, *args, **kwargs)
            self.last_request_time = time.time()
            return result

    def release(self):
        self.manager.release_token(self.token)

# ---------- 解析 Excel 结构（过滤“是否写入”为“是”） ----------
def parse_db_structure(excel_path, force_reload=False):
    """
    返回结构字典：{表名: [字段字典, ...]}，仅包含“是否写入”为“是”的字段。
    同时缓存原始结构供对比用。
    """
    global _DB_STRUCTURE_CACHE, _DB_STRUCTURE_RAW_CACHE
    if _DB_STRUCTURE_CACHE is not None and not force_reload:
        return _DB_STRUCTURE_CACHE

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"数据库结构文件不存在: {excel_path}")

    logger.info("解析 Excel 结构...")
    wb = load_workbook(excel_path, data_only=True)
    structure = {}
    raw_structure = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        # 查找关键列索引
        try:
            idx_en = header.index('变量名称-英文')
        except ValueError:
            idx_en = 2
        try:
            idx_type = header.index('数据类型')
        except ValueError:
            idx_type = 3
        try:
            idx_write = header.index('是否写入')
        except ValueError:
            idx_write = len(header) - 1   # 默认最后一列
        try:
            idx_source = header.index('数据来源')
        except ValueError:
            idx_source = 4
        try:
            idx_desc = header.index('说明')
        except ValueError:
            idx_desc = 5

        fields = []
        raw_fields = []
        seen_names = set()
        for row in rows[1:]:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            en_name = None
            if idx_en < len(row):
                val = row[idx_en]
                if val is not None:
                    en_name = str(val).strip()
            if not en_name:
                continue
            if en_name in seen_names:
                logger.warning(f"表 '{sheet_name}' 重复列名 '{en_name}'，跳过")
                continue
            seen_names.add(en_name)

            field_type = str(row[idx_type]).strip() if idx_type < len(row) else '字符串'
            source = str(row[idx_source]).strip() if idx_source < len(row) else ''
            desc = str(row[idx_desc]).strip() if idx_desc < len(row) else ''
            write_flag = '是'
            if idx_write < len(row):
                write_flag = str(row[idx_write]).strip()

            field_dict = {
                'en': en_name,
                'type': field_type,
                'source': source,
                'desc': desc
            }
            raw_fields.append(field_dict)
            if write_flag == '是':
                fields.append(field_dict)
        structure[sheet_name] = fields
        raw_structure[sheet_name] = raw_fields
        logger.info(f"表 '{sheet_name}' 解析到 {len(fields)} 个有效字段（写入），原始 {len(raw_fields)} 个")

    _DB_STRUCTURE_CACHE = structure
    _DB_STRUCTURE_RAW_CACHE = raw_structure
    return structure

def get_raw_structure(excel_path):
    """获取原始结构（含所有字段，不过滤），用于对比新增字段"""
    global _DB_STRUCTURE_RAW_CACHE
    if _DB_STRUCTURE_RAW_CACHE is None:
        parse_db_structure(excel_path)  # 触发解析
    return _DB_STRUCTURE_RAW_CACHE

def map_type_to_duckdb(field_type):
    t = field_type.strip()
    if '字符串' in t:
        return 'VARCHAR'
    elif '日期' in t:
        return 'DATE'
    elif '整数' in t:
        return 'INTEGER'
    elif '数值' in t or '浮点' in t:
        return 'DOUBLE'
    else:
        return 'VARCHAR'

# ---------- 表结构同步（创建/补充字段） ----------
def sync_table_schema(conn, table_name, fields):
    """
    检查表是否存在，若不存在则创建；若存在则检查字段，补充缺失的字段。
    fields: 列表，每个元素为 {'en': ..., 'type': ...}
    """
    if not fields:
        logger.warning(f"表 '{table_name}' 没有有效字段，跳过")
        return

    # 检查表是否存在
    table_exists = conn.execute(f"SELECT COUNT(*) FROM information_schema.tables WHERE table_name = '{table_name}'").fetchone()[0] > 0
    if not table_exists:
        # 创建表
        cols = []
        for f in fields:
            duck_type = map_type_to_duckdb(f['type'])
            cols.append(f'"{f["en"]}" {duck_type}')
        # 主键逻辑：日度/周度/月度表为 (ts_code, trade_date)，基本/当日表为 ts_code
        if table_name in ['日度信息表', '周度信息表', '月度信息表']:
            pk = ', PRIMARY KEY ("ts_code", "trade_date")'
        elif table_name in ['基本信息表', '当日信息表']:
            pk = ', PRIMARY KEY ("ts_code")'
        else:
            pk = ''
        ddl = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(cols)}{pk})'
        conn.execute(ddl)
        logger.info(f"表 '{table_name}' 创建完成")
    else:
        # 检查现有列
        existing_cols = conn.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}'").fetchall()
        existing_cols = [c[0] for c in existing_cols]
        target_cols = [f['en'] for f in fields]
        missing = [col for col in target_cols if col not in existing_cols]
        if missing:
            # 逐个添加字段
            for col in missing:
                # 从fields中找到类型
                f_type = next((f['type'] for f in fields if f['en'] == col), 'VARCHAR')
                duck_type = map_type_to_duckdb(f_type)
                alter_sql = f'ALTER TABLE "{table_name}" ADD COLUMN "{col}" {duck_type}'
                conn.execute(alter_sql)
                logger.info(f"表 '{table_name}' 添加字段 '{col}'")
        else:
            logger.info(f"表 '{table_name}' 结构已是最新")

# ---------- 数据获取函数 ----------
def fetch_stock_basic(pro):
    df = pro.stock_basic(list_status='L', fields='ts_code,symbol,name,area,industry,fullname,enname,list_date,exchange,is_hs')
    return df if df is not None else pd.DataFrame()

def fetch_company_by_exchange(pro, exchange):
    df = pro.stock_company(exchange=exchange)
    return df if df is not None else pd.DataFrame()

def fetch_daily(pro, trade_date):
    df = pro.daily(trade_date=trade_date)
    return df if df is not None else pd.DataFrame()

def fetch_daily_basic(pro, trade_date):
    df = pro.daily_basic(trade_date=trade_date, fields='ts_code,trade_date,pe,pe_ttm,pb,ps,ps_ttm,total_mv,circ_mv,total_share,float_share,turnover_rate,dv_ratio,dv_ttm,volume_ratio')
    return df if df is not None else pd.DataFrame()

def compute_sector(code):
    code_str = str(code).strip()
    if code_str.startswith('60'):
        return '上证主板'
    elif code_str.startswith('688'):
        return '科创板'
    elif code_str.startswith('00'):
        return '深证主板'
    elif code_str.startswith('30'):
        return '创业板'
    elif code_str.startswith(('92', '83', '87', '43')):
        return '北交所'
    else:
        return '其他'

# ---------- 动态计算最优线程数 ----------
def get_optimal_workers():
    num_tokens = len(TUSHARE_TOKENS)
    optimal = min(num_tokens, 12)
    return max(1, optimal)

# ---------- 核心更新函数 ----------
def update_basic_info(conn, token_manager):
    logger.info("开始更新基本信息表...")
    token = token_manager.acquire_token()
    try:
        df_basic = token.request(fetch_stock_basic)
        if df_basic.empty:
            raise Exception("获取股票基础信息失败")
        exchanges = df_basic['exchange'].unique().tolist()
        df_company_all = pd.DataFrame()
        for ex in exchanges:
            if not ex:
                continue
            logger.info(f"获取交易所 {ex} 的公司信息...")
            df_part = token.request(fetch_company_by_exchange, ex)
            if not df_part.empty:
                df_company_all = pd.concat([df_company_all, df_part], ignore_index=True)
            time.sleep(0.5)

        df_merged = df_basic.merge(df_company_all, on='ts_code', how='left', suffixes=('', '_comp'))
        for col in df_merged.columns:
            if col.endswith('_comp'):
                base = col[:-5]
                if base in df_merged.columns:
                    df_merged[base] = df_merged[base].fillna(df_merged[col])
                    df_merged.drop(columns=[col], inplace=True)

        df_merged['type'] = '股票'
        df_merged['sector'] = df_merged['symbol'].apply(compute_sector)
        for date_col in ['list_date', 'setup_date']:
            if date_col in df_merged.columns:
                df_merged[date_col] = pd.to_datetime(df_merged[date_col], errors='coerce').dt.date

        df_merged = df_merged.where(pd.notna(df_merged), None)

        # 获取目标字段
        structure = parse_db_structure(EXCEL_PATH)
        fields = structure.get('基本信息表', [])
        if not fields:
            raise Exception("基本信息表未定义任何有效字段")
        target_cols = [f['en'] for f in fields]
        for col in target_cols:
            if col not in df_merged.columns:
                df_merged[col] = None
        df_final = df_merged[target_cols]

        # DELETE+INSERT 必须包事务：否则 INSERT 失败会留下空表
        try:
            conn.execute('BEGIN TRANSACTION')
            conn.execute('DELETE FROM "基本信息表"')
            conn.register('df_basic', df_final)
            conn.execute('INSERT INTO "基本信息表" SELECT * FROM df_basic')
            conn.execute('COMMIT')
        except Exception:
            conn.execute('ROLLBACK')
            raise
        logger.info(f"基本信息表更新完成，共 {len(df_final)} 条记录")
    finally:
        token.release()

def update_daily_snapshot(conn, token_manager_normal, token_manager_advanced):
    logger.info("开始更新当日信息表...")
    token = token_manager_normal.acquire_token()
    adv_token = token_manager_advanced.acquire_token() if TUSHARE_TOKENS_2000__ else None
    try:
        pro = ts.pro_api(token.token)
        trade_date = get_latest_trade_date_with_data(pro)
        logger.info(f"最新交易日: {trade_date}")

        # 获取 daily 数据
        df_daily = token.request(fetch_daily, trade_date)
        if df_daily.empty:
            raise Exception(f"获取 daily 行情失败，日期 {trade_date}")

        # 获取 daily_basic 数据（若有高级Token）
        df_basic = pd.DataFrame()
        if adv_token:
            try:
                df_basic = adv_token.request(fetch_daily_basic, trade_date)
                logger.info(f"获取 daily_basic 成功，{len(df_basic)} 条记录")
            except Exception as e:
                logger.warning(f"获取 daily_basic 失败: {e}")
        else:
            logger.warning("未配置2000积分Token，跳过 daily_basic 数据")

        # 合并
        if not df_basic.empty:
            df_merged = df_daily.merge(df_basic, on=['ts_code', 'trade_date'], how='left')
        else:
            df_merged = df_daily.copy()

        # 计算振幅、涨跌幅、涨跌额（若没有则计算）
        if 'amp' not in df_merged.columns:
            df_merged['amp'] = (df_merged['high'] - df_merged['low']) / df_merged['pre_close'] * 100
        if 'pct_chg' not in df_merged.columns:
            df_merged['pct_chg'] = (df_merged['close'] - df_merged['pre_close']) / df_merged['pre_close'] * 100
        if 'change' not in df_merged.columns:
            df_merged['change'] = df_merged['close'] - df_merged['pre_close']

        df_merged = df_merged.where(pd.notna(df_merged), None)

        # 获取目标字段
        structure = parse_db_structure(EXCEL_PATH)
        fields = structure.get('当日信息表', [])
        if not fields:
            raise Exception("当日信息表未定义任何有效字段")
        target_cols = [f['en'] for f in fields]
        for col in target_cols:
            if col not in df_merged.columns:
                df_merged[col] = None
        df_final = df_merged[target_cols]
        df_final['trade_date'] = pd.to_datetime(df_final['trade_date'], format='%Y%m%d').dt.date

        # DELETE+INSERT 必须包事务：否则 INSERT 失败会留下空表
        try:
            conn.execute('BEGIN TRANSACTION')
            conn.execute('DELETE FROM "当日信息表"')
            conn.register('df_snapshot', df_final)
            conn.execute('INSERT INTO "当日信息表" SELECT * FROM df_snapshot')
            conn.execute('COMMIT')
        except Exception:
            conn.execute('ROLLBACK')
            raise
        logger.info(f"当日信息表更新完成，共 {len(df_final)} 条记录")
    finally:
        token.release()
        if adv_token:
            adv_token.release()

def update_daily_history(conn, token_manager, incremental=True):
    """
    更新日度信息表，支持增量/全量。
    incremental=True 时只获取数据库中已有最大日期之后的交易日。
    """
    logger.info("开始更新日度信息表...")

    # 1. 获取交易日列表
    token = token_manager.acquire_token()
    try:
        pro = ts.pro_api(token.token)
        all_dates = get_trade_calendar(pro)
        if not all_dates:
            raise Exception("无法获取交易日历")

        if incremental:
            # 检查日度表是否有数据
            max_date_df = conn.execute('SELECT MAX(trade_date) FROM "日度信息表"').fetchone()
            existing_max = max_date_df[0] if max_date_df and max_date_df[0] else None
            if existing_max:
                # 只获取 existing_max 之后的交易日
                all_dates = [d for d in all_dates if d > existing_max.strftime('%Y%m%d')]
                if not all_dates:
                    logger.info("日度信息表已是最新，无需更新")
                    return
                logger.info(f"增量更新，将获取 {len(all_dates)} 个交易日")
            else:
                logger.info("日度信息表为空，执行全量更新")
                conn.execute('DELETE FROM "日度信息表"')
                incremental = False   # 改为全量
        else:
            logger.info("全量更新日度信息表，将清空现有数据")
            conn.execute('DELETE FROM "日度信息表"')
    finally:
        token.release()

    # 2. 数据处理和入库函数
    def process_and_insert(df_daily, date_str):
        try:
            if df_daily.empty:
                return False
            df_merged = df_daily.copy()
            # 计算指标
            df_merged['amp'] = (df_merged['high'] - df_merged['low']) / df_merged['pre_close'] * 100
            df_merged['pct_chg'] = (df_merged['close'] - df_merged['pre_close']) / df_merged['pre_close'] * 100
            df_merged['change'] = df_merged['close'] - df_merged['pre_close']
            df_merged = df_merged.where(pd.notna(df_merged), None)

            structure = parse_db_structure(EXCEL_PATH)
            fields = structure.get('日度信息表', [])
            if not fields:
                raise Exception("日度信息表未定义任何有效字段")
            target_cols = [f['en'] for f in fields]
            for col in target_cols:
                if col not in df_merged.columns:
                    df_merged[col] = None
            df_final = df_merged[target_cols]
            df_final['trade_date'] = pd.to_datetime(df_final['trade_date'], format='%Y%m%d').dt.date

            conn_local = get_thread_db_connection()
            conn_local.register('df_ins', df_final)
            col_names = ', '.join([f'"{col}"' for col in target_cols])
            sql = f"""
                INSERT INTO "日度信息表" ({col_names})
                SELECT * FROM df_ins
                ON CONFLICT ("ts_code", "trade_date") DO NOTHING
            """
            conn_local.execute(sql)
            logger.info(f"日期 {date_str} 处理完成，插入 {len(df_final)} 条")
            return True
        except Exception as e:
            logger.error(f"日期 {date_str} 数据处理入库失败: {e}")
            return False

    # 3. 单个日期的处理流程
    def process_single_date(date_str, token_obj):
        try:
            df_daily = token_obj.request(fetch_daily, date_str)
        except Exception as e:
            error_msg = str(e)
            if '频率超限' in error_msg or '频率限制' in error_msg:
                error_type = 'frequency'
            else:
                error_type = 'other'
            return False, error_type
        finally:
            token_obj.release()

        if df_daily.empty:
            return False, 'empty'

        success = process_and_insert(df_daily, date_str)
        return success, None if success else 'db_error'

    # 4. 并发处理一批日期
    def process_batch(dates, max_workers):
        if not dates:
            return []
        failed = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for date_str in dates:
                token_obj = token_manager.acquire_token()
                future = executor.submit(process_single_date, date_str, token_obj)
                futures[future] = date_str
            for future in as_completed(futures):
                date_str = futures[future]
                try:
                    success, error_type = future.result(timeout=60)
                    if not success:
                        failed.append((date_str, error_type))
                except Exception as e:
                    logger.error(f"日期 {date_str} 线程执行异常: {e}")
                    failed.append((date_str, 'thread_exception'))
        return failed

    # 5. 执行多轮并发
    optimal_workers = get_optimal_workers()
    logger.info(f"并发线程数: {optimal_workers}")

    failed_dates = process_batch(all_dates, optimal_workers)
    logger.info(f"第一轮完成，失败 {len(failed_dates)} 个日期")

    # 重试最多5轮
    for round_num in range(5):
        if not failed_dates:
            break
        retry_list = [date for date, _ in failed_dates]
        logger.info(f"第 {round_num+1} 轮重试，剩余 {len(retry_list)} 个日期")
        new_failed = process_batch(retry_list, optimal_workers)
        failed_dates = new_failed
        logger.info(f"第 {round_num+1} 轮重试完成，剩余 {len(failed_dates)} 个日期")

    # 最终降级：单线程间隔1.5秒
    if failed_dates:
        logger.warning(f"5轮重试后仍剩余 {len(failed_dates)} 个日期，启动最终单线程降级（间隔1.5秒）")
        single_token = token_manager.acquire_token()
        try:
            last_req_time = 0
            interval = 1.5
            success_count = 0
            for date_str, error_type in failed_dates:
                now = time.time()
                wait = interval - (now - last_req_time)
                if wait > 0:
                    time.sleep(wait)
                try:
                    ts.set_token(single_token.token)
                    pro = ts.pro_api()
                    df_daily = pro.daily(trade_date=date_str)
                    last_req_time = time.time()
                    if df_daily is not None and not df_daily.empty:
                        if process_and_insert(df_daily, date_str):
                            success_count += 1
                            logger.info(f"最终降级成功处理日期 {date_str}")
                        else:
                            logger.warning(f"最终降级日期 {date_str} 入库失败")
                    else:
                        logger.warning(f"最终降级日期 {date_str} 无数据")
                except Exception as e:
                    logger.error(f"最终降级日期 {date_str} 请求失败: {e}")
            logger.info(f"最终降级完成，成功 {success_count} 个，失败 {len(failed_dates)-success_count} 个")
        finally:
            single_token.release()

    logger.info("日度信息表更新完成")

def update_weekly_monthly(conn):
    logger.info("开始重建周度/月度信息表...")
    structure = parse_db_structure(EXCEL_PATH)
    weekly_fields = [f['en'] for f in structure.get('周度信息表', [])]
    monthly_fields = [f['en'] for f in structure.get('月度信息表', [])]

    # 周度聚合
    weekly_select_cols = [
        'w.ts_code', 'w.last_trade_date AS trade_date',
        'w.open_w', 'w.close_w', 'w.high_w', 'w.low_w',
        'w.vol_w', 'w.amount_w', 'w.pre_close_w',
        '(w.high_w - w.low_w) / NULLIF(w.pre_close_w, 0) * 100 AS amp_w',
        '(w.close_w - w.pre_close_w) / NULLIF(w.pre_close_w, 0) * 100 AS pct_chg_w',
        'w.close_w - w.pre_close_w AS change_w'
    ]
    # 如果周度表需要 turnover_w，则补 NULL（日度表无 turnover）
    if 'turnover_w' in weekly_fields:
        weekly_select_cols.append('NULL AS turnover_w')
    weekly_sql = f"""
    WITH weekly_agg AS (
        SELECT ts_code, date_trunc('week', trade_date) AS week_start,
               MIN(trade_date) AS first_trade_date, MAX(trade_date) AS last_trade_date,
               FIRST(open ORDER BY trade_date) AS open_w,
               LAST(close ORDER BY trade_date) AS close_w,
               MAX(high) AS high_w, MIN(low) AS low_w,
               SUM(vol) AS vol_w, SUM(amount) AS amount_w
        FROM "日度信息表"
        GROUP BY ts_code, date_trunc('week', trade_date)
    ),
    weekly_with_prev AS (
        SELECT *, LAG(close_w) OVER (PARTITION BY ts_code ORDER BY last_trade_date) AS pre_close_w
        FROM weekly_agg
    )
    SELECT {', '.join(weekly_select_cols)}
    FROM weekly_with_prev w WHERE w.pre_close_w IS NOT NULL
    """
    monthly_select_cols = [
        'm.ts_code', 'm.last_trade_date AS trade_date',
        'm.open_m', 'm.close_m', 'm.high_m', 'm.low_m',
        'm.vol_m', 'm.amount_m', 'm.pre_close_m',
        '(m.high_m - m.low_m) / NULLIF(m.pre_close_m, 0) * 100 AS amp_m',
        '(m.close_m - m.pre_close_m) / NULLIF(m.pre_close_m, 0) * 100 AS pct_chg_m',
        'm.close_m - m.pre_close_m AS change_m'
    ]
    if 'turnover_m' in monthly_fields:
        monthly_select_cols.append('NULL AS turnover_m')
    monthly_sql = f"""
    WITH monthly_agg AS (
        SELECT ts_code, date_trunc('month', trade_date) AS month_start,
               MIN(trade_date) AS first_trade_date, MAX(trade_date) AS last_trade_date,
               FIRST(open ORDER BY trade_date) AS open_m,
               LAST(close ORDER BY trade_date) AS close_m,
               MAX(high) AS high_m, MIN(low) AS low_m,
               SUM(vol) AS vol_m, SUM(amount) AS amount_m
        FROM "日度信息表"
        GROUP BY ts_code, date_trunc('month', trade_date)
    ),
    monthly_with_prev AS (
        SELECT *, LAG(close_m) OVER (PARTITION BY ts_code ORDER BY last_trade_date) AS pre_close_m
        FROM monthly_agg
    )
    SELECT {', '.join(monthly_select_cols)}
    FROM monthly_with_prev m WHERE m.pre_close_m IS NOT NULL
    """
    # DELETE+INSERT 必须包事务：任一 INSERT 失败则回滚，避免两表数据不一致或留空表
    try:
        conn.execute('BEGIN TRANSACTION')
        conn.execute('DELETE FROM "周度信息表"')
        conn.execute('DELETE FROM "月度信息表"')
        conn.execute('INSERT INTO "周度信息表" ' + weekly_sql)
        logger.info("周度信息表重建完成")
        conn.execute('INSERT INTO "月度信息表" ' + monthly_sql)
        logger.info("月度信息表重建完成")
        conn.execute('COMMIT')
    except Exception:
        conn.execute('ROLLBACK')
        raise

# ---------- 统一更新入口 ----------
def full_update():
    """全量更新：清空所有数据，重新导入全部历史"""
    logger.info("========== 开始全量更新 ==========")
    start_time = time.time()
    os.makedirs(INSTANCE_DIR, exist_ok=True)

    # 解析结构
    structure = parse_db_structure(EXCEL_PATH)
    conn = duckdb.connect(DB_PATH)
    # 创建所有表（确保结构）
    for table_name, fields in structure.items():
        sync_table_schema(conn, table_name, fields)
    conn.close()

    token_manager = TokenManager(TUSHARE_TOKENS)
    adv_token_manager = TokenManager(TUSHARE_TOKENS_2000__) if TUSHARE_TOKENS_2000__ else None

    conn = duckdb.connect(DB_PATH)
    try:
        update_basic_info(conn, token_manager)
        update_daily_snapshot(conn, token_manager, adv_token_manager)
        update_daily_history(conn, token_manager, incremental=False)
        update_weekly_monthly(conn)
    finally:
        conn.close()

    elapsed = time.time() - start_time
    logger.info(f"全量更新完成，总耗时 {elapsed:.2f} 秒")

def incremental_update():
    """
    增量更新（智能更新）
    1. 检查 Excel 结构文件是否存在，否则抛出 FileNotFoundError
    2. 检查数据库是否存在，若不存在则执行全量更新
    3. 若存在，则同步表结构（创建/添加字段）
    4. 更新基本信息表（全量替换）、当日信息表（全量替换）
    5. 更新日度信息表（增量）
    6. 重建周度/月度信息表
    """
    # 检查结构文件
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"缺少数据库结构文件: {EXCEL_PATH}")

    logger.info("========== 开始增量更新 ==========")
    start_time = time.time()
    os.makedirs(INSTANCE_DIR, exist_ok=True)

    # 解析结构（仅有效字段）
    structure = parse_db_structure(EXCEL_PATH)
    conn = duckdb.connect(DB_PATH)

    # 同步所有表结构
    for table_name, fields in structure.items():
        sync_table_schema(conn, table_name, fields)

    # 判断日度表是否有数据，决定全量还是增量
    has_daily_data = conn.execute('SELECT COUNT(*) FROM "日度信息表"').fetchone()[0] > 0

    token_manager = TokenManager(TUSHARE_TOKENS)
    adv_token_manager = TokenManager(TUSHARE_TOKENS_2000__) if TUSHARE_TOKENS_2000__ else None

    try:
        # 更新基本信息表（全量替换）
        update_basic_info(conn, token_manager)

        # 更新当日信息表（全量替换）
        update_daily_snapshot(conn, token_manager, adv_token_manager)

        # 更新日度历史（全量或增量）
        update_daily_history(conn, token_manager, incremental=has_daily_data)

        # 重建周/月度
        update_weekly_monthly(conn)
    finally:
        conn.close()

    elapsed = time.time() - start_time
    logger.info(f"增量更新完成，总耗时 {elapsed:.2f} 秒")

if __name__ == '__main__':
    # 直接运行脚本时执行全量更新
    full_update()
